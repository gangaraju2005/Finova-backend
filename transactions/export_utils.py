import csv
from io import StringIO, BytesIO
from datetime import datetime
from django.utils.dateparse import parse_datetime
"""
transactions/export_utils.py

Utilities for exporting and importing transaction data.

Export Columns: Date, Description, Category, Type (INCOME/EXPENSE), Amount, Payment Method
Import Columns: date, description, category, type, amount, payment_method
"""
import csv
from io import StringIO, BytesIO
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXPORT_HEADERS = ['Date', 'Description', 'Category', 'Type', 'Amount', 'Payment Method']


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def _tx_row(tx):
    """Convert a Transaction object to a list of values for export."""
    return [
        tx.date.strftime('%Y-%m-%d') if tx.date else '',
        tx.description or '',
        tx.category.name if tx.category else 'Uncategorized',
        tx.category.type if tx.category else 'EXPENSE',   # Raw: INCOME or EXPENSE
        float(tx.amount),
        tx.payment_method or 'CASH',
    ]


def generate_csv_export(transactions):
    """Returns a CSV string for all given transactions."""
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['Date', 'Description', 'Category', 'Type', 'Amount', 'Payment Method'])
    
    for tx in transactions:
        writer.writerow([
            tx.date.strftime('%Y-%m-%d'),
            tx.description,
            tx.category.name if tx.category else 'Uncategorized',
            tx.category.get_type_display() if tx.category else 'Expense',
            tx.amount,
            tx.payment_method
        ])
    
    return output.getvalue().encode('utf-8')
    writer.writerow(EXPORT_HEADERS)
    for tx in transactions:
        writer.writerow(_tx_row(tx))
    return output.getvalue()


def generate_xlsx_export(transactions):
    """Returns bytes of an xlsx workbook for all given transactions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Headers
    headers = ['Date', 'Description', 'Category', 'Type', 'Amount', 'Payment Method']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:

        # ── Header row styling ──────────────────────────────────────────────────
        header_fill = PatternFill(start_color='1A1F36', end_color='1A1F36', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center')

    ws.append(EXPORT_HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────
    for tx in transactions:
        ws.append([
            tx.date.strftime('%Y-%m-%d'),
            tx.description,
            tx.category.name if tx.category else 'Uncategorized',
            tx.category.get_type_display() if tx.category else 'Expense',
            tx.amount,
            tx.payment_method
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.append(_tx_row(tx))

    # ── Auto-adjust column widths ──────────────────────────────────────────
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_sample_csv():
    """Generates a sample CSV template for users to follow during import."""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['date', 'amount', 'category', 'type', 'payment_method', 'description'])
    writer.writerow(['2024-03-24', '150.00', 'Groceries', 'EXPENSE', 'CREDIT_CARD', 'Weekly groceries'])
    writer.writerow(['2024-03-25', '2500.00', 'Salary', 'INCOME', 'BANK_TRANSFER', 'Monthly paycheck'])
    return output.getvalue().encode('utf-8')

def parse_csv_import(file_file):
    import csv

# ─────────────────────────────────────────────────────────────────────────────
# IMPORT
# ─────────────────────────────────────────────────────────────────────────────
# Expected import columns (case-insensitive):
#   date, description, category, type, amount, payment_method

def _normalize_row(raw: dict) -> dict:
    """
    Normalize a raw dict from CSV/XLSX into a clean transaction dict.
    Returns the dict with standardised keys and a list of validation errors.
    """
    errors = []

    # ── date ──
    date_str = str(raw.get('date') or raw.get('Date') or '').strip()
    tx_date = None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            tx_date = datetime.strptime(date_str.split(' ')[0].split('T')[0], fmt)
            break
        except (ValueError, AttributeError):
            pass
    if not tx_date:
        if date_str:
            errors.append(f"Unrecognised date format: '{date_str}'")
        tx_date = datetime.now()

    # ── amount ──
    raw_amount = str(raw.get('amount') or raw.get('Amount') or '').strip().replace(',', '')
    try:
        amount = float(raw_amount)
        if amount <= 0:
            errors.append("Amount must be greater than zero")
    except ValueError:
        errors.append(f"Invalid amount: '{raw_amount}'")
        amount = None

    # ── type ──
    raw_type = str(raw.get('type') or raw.get('Type') or 'EXPENSE').strip().upper()
    if raw_type not in ('INCOME', 'EXPENSE'):
        raw_type = 'EXPENSE'

    # ── category ──
    category_name = str(raw.get('category') or raw.get('Category') or 'Uncategorized').strip()

    # ── description ──
    description = str(raw.get('description') or raw.get('Description') or '').strip()
    if not description:
        errors.append("Missing description")

    # ── payment method ──
    valid_methods = ('CASH', 'CARD', 'UPI', 'BANK_TRANSFER', 'CHEQUE', 'OTHER')
    payment_method = str(raw.get('payment_method') or raw.get('Payment Method') or 'CASH').strip().upper().replace(' ', '_')
    if payment_method not in valid_methods:
        payment_method = 'CASH'

    return {
        'date': tx_date,
        'description': description,
        'category_name': category_name,
        'category_type': raw_type,
        'amount': amount,
        'payment_method': payment_method,
        'errors': errors,
    }


def parse_csv_import(file_obj):
    """Parse a CSV file object and return a list of normalised row dicts."""
    from io import TextIOWrapper
    try:
        text_file = TextIOWrapper(file_obj, encoding='utf-8-sig')
    except TypeError:
        text_file = file_obj

    reader = csv.DictReader(text_file)
    
    transactions_data = []
    for row in reader:
        # Normalise keys to lower for easier parsing
        data = {k.lower().strip(): v for k, v in row.items()}
        transactions_data.append({
            'date': data.get('date'),
            'description': data.get('description'),
            'category_name': data.get('category'),
            'category_type': data.get('type'),
            'amount': data.get('amount'),
            'payment_method': data.get('payment_method')
        })
    return transactions_data

def parse_xlsx_import(file_file):
    wb = load_workbook(file_file, data_only=True)
    ws = wb.active
    
    transactions_data = []
    rows = list(ws.rows)
    if not rows:
        return []
        
    headers = [str(cell.value).lower().strip() for cell in rows[0]]
    
    for row in rows[1:]:
        row_data = {headers[i]: row[i].value for i in range(len(headers)) if i < len(headers)}
        transactions_data.append({
            'date': str(row_data.get('date')) if row_data.get('date') else None,
            'description': row_data.get('description'),
            'category_name': row_data.get('category'),
            'category_type': row_data.get('type'),
            'amount': row_data.get('amount'),
            'payment_method': row_data.get('payment_method')
        })
    return transactions_data
    rows = []
    for i, row in enumerate(reader, start=2):     # row 1 = header
        norm = _normalize_row(dict(row))
        norm['row_num'] = i
        rows.append(norm)
    return rows


def parse_xlsx_import(file_obj):
    """Parse an XLSX file object and return a list of normalised row dicts."""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    raw_rows = list(ws.rows)
    if not raw_rows:
        return []

    headers = [str(cell.value).strip() if cell.value else '' for cell in raw_rows[0]]
    rows = []
    for i, row in enumerate(raw_rows[1:], start=2):
        raw = {headers[j]: (row[j].value if j < len(row) else None) for j in range(len(headers))}
        norm = _normalize_row(raw)
        norm['row_num'] = i
        rows.append(norm)
    return rows


def generate_sample_csv():
    """Returns a sample CSV string users can fill in for importing."""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    writer.writerow(['2024-01-15', 'Grocery Shopping', 'Food', 'EXPENSE', 450.00, 'CARD'])
    writer.writerow(['2024-01-16', 'Salary Credit', 'Salary', 'INCOME', 50000.00, 'BANK_TRANSFER'])
    writer.writerow(['2024-01-17', 'Netflix', 'Entertainment', 'EXPENSE', 649.00, 'CARD'])
    return output.getvalue()
